"""Phase 2C step 1: extract per-entry activity labels from DRAMP metadata.

DRAMP's bulk FASTA carries only entry IDs (e.g. `>DRAMP00005`); the
activity annotations live in a separate XLSX/TSV metadata file. This
script tries to download that metadata, joins it with amps.fasta by
DRAMP ID, and writes a per-entry multi-label activity CSV.

Maps DRAMP's free-text activity strings into four canonical Phase 2C
classes:

  - antibacterial  ← Antibacterial, Anti-Gram+, Anti-Gram-, Antibiofilm
  - antifungal     ← Antifungal
  - antiviral      ← Antiviral, Anti-HIV
  - antiparasitic  ← Antiparasitic, Antiprotozoal

Each peptide can have multiple positive labels (multi-label, not
mutually exclusive). DRAMP entries with no recognised activity in
the metadata get all four columns set to 0 and an `other` flag set to
1, so the downstream trainer can drop them.

Saves: ml/data/raw/amp_activities.csv with columns
  dramp_id, sequence, antibacterial, antifungal, antiviral,
  antiparasitic, other, raw_activity
"""
from __future__ import annotations

import io
import re
import sys
import time
import zipfile
from pathlib import Path

import requests

PROJECT_ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = PROJECT_ROOT / "ml" / "data" / "raw"
AMPS_FASTA = RAW_DIR / "amps.fasta"
METADATA_PATH = RAW_DIR / "dramp_general_metadata.xlsx"
ACTIVITIES_CSV = RAW_DIR / "amp_activities.csv"
MISSING_PATH = RAW_DIR / "MISSING_ACTIVITIES.md"

USER_AGENT = "AMPClassifier/0.1 (research; portfolio project)"
TIMEOUT = 30

# DRAMP metadata download candidates (XLSX or TSV).
METADATA_URLS = (
    "https://dramp.cpu-bioinfor.org/static/downloads/general_amps.xlsx",
    "https://dramp.cpu-bioinfor.org/static/downloads/General_amps.xlsx",
    "https://dramp.cpu-bioinfor.org/static/downloads/general_amps_v4.xlsx",
    "https://dramp.cpu-bioinfor.org/static/downloads/general_amps_v3.xlsx",
    "https://dramp.cpu-bioinfor.org/static/downloads/general_amps.txt",
    "https://dramp.cpu-bioinfor.org/static/downloads/general_amps.tsv",
    "https://dramp.cpu-bioinfor.org/static/downloads/general_amps.csv",
)

INDEX_URLS = (
    "https://dramp.cpu-bioinfor.org/downloads/",
    "http://dramp.cpu-bioinfor.org/downloads/",
)

# Mapping from DRAMP activity strings (lowercased, substring match) to
# our four canonical classes. A peptide can hit multiple classes.
ACTIVITY_MAP = {
    "antibacterial":  ("antibacterial", "anti-gram", "antibiofilm", "antimicrob",
                        "bactericid", "anti-bacter"),
    "antifungal":     ("antifungal", "anti-fungal"),
    "antiviral":      ("antiviral", "anti-viral", "anti-hiv", "anti-hcv", "anti-sars"),
    "antiparasitic":  ("antiparasit", "antiprotozo", "anti-malarial", "antimalarial",
                        "anti-leishman", "antitrypano"),
}


# ---------------------------------------------------------------------------
# HTTP helpers (small subset of download_data.py helpers)
# ---------------------------------------------------------------------------


def _session():
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT})
    return s


def fetch_metadata(session):
    """Try documented URLs then scrape the /downloads/ index for an XLSX/TSV."""
    for url in METADATA_URLS:
        print(f"[activities] trying {url}")
        try:
            r = session.get(url, timeout=TIMEOUT)
            if r.status_code == 200 and len(r.content) > 1024:
                ext = url.rsplit(".", 1)[-1].lower()
                return r.content, ext
        except requests.RequestException as exc:
            print(f"  [activities] {exc!r}")

    # Index scrape — look for any .xlsx, .tsv, or .csv with "general" in path
    pat = re.compile(r'href=["\']([^"\']+\.(?:xlsx|tsv|csv|txt))["\']', re.IGNORECASE)
    seen = set()
    candidates = []
    for index_url in INDEX_URLS:
        try:
            r = session.get(index_url, timeout=TIMEOUT)
        except requests.RequestException:
            continue
        if r.status_code != 200:
            continue
        for raw in pat.findall(r.text):
            from urllib.parse import urljoin
            absu = urljoin(index_url, raw)
            if absu in seen:
                continue
            seen.add(absu)
            score = 0
            if "general" in absu.lower():
                score += 10
            if absu.lower().endswith(".xlsx"):
                score += 3
            elif absu.lower().endswith(".tsv"):
                score += 2
            elif absu.lower().endswith(".csv"):
                score += 1
            candidates.append((score, absu))
    candidates.sort(reverse=True)
    for _, url in candidates:
        print(f"[activities] trying (index) {url}")
        try:
            r = session.get(url, timeout=TIMEOUT)
            if r.status_code == 200 and len(r.content) > 1024:
                ext = url.rsplit(".", 1)[-1].lower()
                return r.content, ext
        except requests.RequestException:
            pass
    return None, None


# ---------------------------------------------------------------------------
# Parse the metadata content into a dict[dramp_id] -> raw activity string
# ---------------------------------------------------------------------------


def parse_xlsx(content):
    """Parse XLSX bytes via openpyxl. Returns list of dicts (row dicts)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    out = []
    for row in rows_iter:
        record = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
        out.append(record)
    return out, header


def parse_tabular(content, delim):
    """Parse TSV/CSV bytes. Returns list of row dicts + header."""
    import csv
    text = content.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    header = reader.fieldnames or []
    return list(reader), [h.strip() for h in header]


def find_columns(header):
    """Pick the ID column and the activity column from a noisy header list."""
    id_col = activity_col = None
    for h in header:
        hl = h.lower()
        if id_col is None and ("dramp" in hl and "id" in hl):
            id_col = h
        if id_col is None and hl in ("dramp_id", "drampid", "id"):
            id_col = h
        if activity_col is None and "activit" in hl:
            activity_col = h
    return id_col, activity_col


def map_activities(raw_str):
    """Return dict of canonical class -> 0/1 plus a True/False 'other' flag."""
    text = (raw_str or "").lower()
    flags = {cls: 0 for cls in ACTIVITY_MAP}
    matched = False
    for cls, keys in ACTIVITY_MAP.items():
        if any(k in text for k in keys):
            flags[cls] = 1
            matched = True
    return flags, (not matched)


# ---------------------------------------------------------------------------
# FASTA parse + main
# ---------------------------------------------------------------------------


def parse_fasta(path):
    header, chunks = None, []
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.rstrip()
            if line.startswith(">"):
                if header is not None:
                    yield header, "".join(chunks)
                header = line[1:].strip().split()[0]
                chunks = []
            elif line.strip():
                chunks.append(line.strip())
    if header is not None:
        yield header, "".join(chunks)


def write_missing(reason):
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    MISSING_PATH.write_text(
        "# DRAMP activity metadata acquisition failed\n\n"
        f"{reason}\n\n"
        "Manual fix:\n\n"
        "1. Open https://dramp.cpu-bioinfor.org/downloads/ in a browser.\n"
        "2. Find the 'General AMPs' download row, click the XLSX or TSV link.\n"
        "3. Save the file as `ml/data/raw/dramp_general_metadata.xlsx`\n"
        "   (or .tsv / .csv — the script auto-detects).\n"
        "4. Re-run `python ml/scripts/extract_amp_activities.py`.\n",
        encoding="utf-8",
    )
    print(f"[activities] wrote {MISSING_PATH}")


def main():
    if not AMPS_FASTA.exists():
        print(f"[activities] {AMPS_FASTA} missing. Run dev.bat data first.")
        return 1
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    # Step 1 — get metadata content (download or cached)
    content = ext = None
    if METADATA_PATH.exists() and METADATA_PATH.stat().st_size > 1024:
        print(f"[activities] using cached {METADATA_PATH.name}")
        content = METADATA_PATH.read_bytes()
        ext = "xlsx"
    else:
        session = _session()
        content, ext = fetch_metadata(session)
        if content is None:
            write_missing("DRAMP metadata download failed via documented URLs and "
                          "index scrape.")
            return 1
        out_path = RAW_DIR / f"dramp_general_metadata.{ext}"
        out_path.write_bytes(content)
        print(f"[activities] saved {out_path.name} ({len(content):,} bytes)")
        if ext != "xlsx":
            # Keep the canonical filename pointing at XLSX too if applicable
            pass

    # Step 2 — parse
    print(f"[activities] parsing as {ext}")
    if ext == "xlsx":
        try:
            from openpyxl import load_workbook  # noqa: F401
        except ImportError:
            print("[activities] openpyxl not installed; "
                  "run `pip install openpyxl>=3.0`")
            return 1
        rows, header = parse_xlsx(content)
    elif ext == "tsv":
        rows, header = parse_tabular(content, "\t")
    elif ext in ("csv", "txt"):
        # txt could be tab- or comma-separated; sniff once
        sample = content[:2048].decode("utf-8", errors="replace")
        delim = "\t" if sample.count("\t") > sample.count(",") else ","
        rows, header = parse_tabular(content, delim)
    else:
        write_missing(f"Unknown metadata file extension: {ext}")
        return 1

    id_col, activity_col = find_columns(header)
    if id_col is None or activity_col is None:
        write_missing(
            f"Could not auto-detect DRAMP ID and activity columns in metadata.\n"
            f"Headers found: {header}\n"
            f"Detected: id_col={id_col}, activity_col={activity_col}\n\n"
            f"Edit `find_columns` in this script to handle the actual column names."
        )
        return 1
    print(f"[activities] using id_col='{id_col}', activity_col='{activity_col}'")

    # Build dramp_id -> activity raw string
    meta_lookup = {}
    for row in rows:
        rid = str(row.get(id_col, "") or "").strip()
        if rid:
            meta_lookup[rid] = str(row.get(activity_col, "") or "")
    print(f"[activities] metadata covers {len(meta_lookup)} DRAMP entries")

    # Step 3 — join with amps.fasta, write CSV
    import csv
    n_total = 0
    n_matched = 0
    class_counts = {c: 0 for c in ACTIVITY_MAP}
    other_count = 0
    with ACTIVITIES_CSV.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["dramp_id", "sequence", "antibacterial", "antifungal",
                    "antiviral", "antiparasitic", "other", "raw_activity"])
        for header_line, seq in parse_fasta(AMPS_FASTA):
            n_total += 1
            rid = header_line.split()[0].strip()
            raw_act = meta_lookup.get(rid, "")
            flags, is_other = map_activities(raw_act)
            if raw_act:
                n_matched += 1
            for c, v in flags.items():
                if v:
                    class_counts[c] += 1
            if is_other:
                other_count += 1
            w.writerow([rid, seq,
                        flags["antibacterial"], flags["antifungal"],
                        flags["antiviral"], flags["antiparasitic"],
                        1 if is_other else 0, raw_act])

    print(f"[activities] wrote {ACTIVITIES_CSV.name}: {n_total} positives, "
          f"{n_matched} matched in metadata")
    print(f"[activities] class counts (multi-label, can sum > total):")
    for c, n in class_counts.items():
        print(f"  {c}: {n} ({100*n/max(1,n_total):.1f}%)")
    print(f"  other_only: {other_count} ({100*other_count/max(1,n_total):.1f}%)")
    if MISSING_PATH.exists():
        MISSING_PATH.unlink()
    return 0


if __name__ == "__main__":
    sys.exit(main())
